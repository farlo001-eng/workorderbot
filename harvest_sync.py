"""
harvest_sync.py
===============
Downloads and triages open work orders from Yardi:
  1. Log into Yardi via Selenium, download open work orders as .xls
  2. Run AI triage (GPT-4o-mini) to assign AI Priority + AI Reason
  3. Save enriched data to I:\PycharmProjects\WorkOrderBot\WorkOrders.xlsx

Does NOT sync to Railway. Run sync_workorders.py separately afterward to
POST the saved WorkOrders.xlsx to the Railway sync endpoint.

Run with: py harvest_sync.py
"""

import os
import re
import glob
import json
import time
import shutil
import openpyxl
import pandas as pd

from datetime import datetime, timedelta
from dotenv import load_dotenv
from openai import OpenAI
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
import win32com.client
import subprocess

# ---------- CONFIG ----------

load_dotenv()

OPENAI_API_KEY  = os.getenv("OPENAI_API_KEY")

# Yardi credentials from yardi.xlsx (existing approach)
YARDI_CREDS_PATH = r"I:\PycharmProjects\yardi.xlsx"

# Where Yardi downloads land
DOWNLOAD_DIR    = r"G:\Harvest Apartment Management\Work Orders\Data"

# Final output path (read separately by sync_workorders.py)
OUTPUT_PATH     = r"I:\PycharmProjects\WorkOrderBot\WorkOrders.xlsx"
OUTPUT_SHEET    = "Open WO Data"

OPENAI_MODEL    = "gpt-4o-mini"
MAX_ROWS_FOR_AI = 500

# Property/tech lookup table, loaded once in main()
PROPERTIES_XLSX_PATH = r"I:\PycharmProjects\WorkOrderBot\Properties.xlsx"
PROPERTIES_MAP = {}

EMERGENCY_KEYWORDS = [
    "smell gas", "gas leak", "smells like gas", "carbon monoxide",
    "fire", "smoke", "sparks", "burning smell",
    "water pouring", "ceiling collapsing", "ceiling collapse",
    "flooding", "flooded",
]

RISKY_WORDS = [
    "no heat", "heat not working", "heater not working",
    "no ac", "no a/c", "ac not working", "air not working",
    "no water", "no running water",
    "sewage", "sewer backup", "sewage backup",
]

# ---------- PROPERTIES LOOKUP ----------

def normalize_code(val):
    if val is None:
        return ""
    try:
        return str(int(float(str(val).strip())))
    except Exception:
        return str(val).strip()


def load_properties_map(path: str) -> dict:
    """
    Returns dict keyed by normalized property code string.
    E.g. {"79800": {"property_name": "CityCenter Place", "tech_name": "Antonio Smith"}, ...}
    """
    df = pd.read_excel(path, dtype=str)
    result = {}
    for _, row in df.iterrows():
        code = normalize_code(row.get("Property", ""))
        if not code:
            continue
        result[code] = {
            "property_name": str(row.get("Property Name", "") or "").strip(),
            "tech_name":     str(row.get("Maintenance Tech", "") or "").strip(),
        }
    return result


# ---------- STEP 1: YARDI DOWNLOAD ----------

def clear_download_dir():
    for filename in os.listdir(DOWNLOAD_DIR):
        file_path = os.path.join(DOWNLOAD_DIR, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print(f"Warning: could not delete {file_path}: {e}")


def get_yardi_credentials():
    wb = openpyxl.load_workbook(YARDI_CREDS_PATH)
    ws = wb['Sheet1']
    username = ws.cell(row=1, column=1).value
    password = ws.cell(row=2, column=1).value
    wb.close()
    return username, password


def get_mfa_code():
    """Read MFA code from Outlook inbox."""
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6)
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)
    for message in messages:
        if "no-reply@yardione.com" in message.SenderEmailAddress.lower():
            match = re.search(r'Your YardiOne code is:\s*(\d{6})', message.Body)
            if match:
                return match.group(1)
    return None


def download_open_wo():
    """
    Log into Yardi, navigate to Work Orders, download open WOs as .xls.
    Returns the path to the downloaded .xls file.
    """
    today     = datetime.today().strftime('%m/%d/%Y')
    username, password = get_yardi_credentials()

    clear_download_dir()

    chrome_options = Options()
    chrome_options.add_experimental_option("prefs", {
        "download.default_directory": DOWNLOAD_DIR
    })
    chrome_options.add_argument("--enable-javascript")
    # chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--window-size=1920,1080")

    chrome_service = ChromeService(ChromeDriverManager().install())
    chrome_service.creationflags = subprocess.CREATE_NO_WINDOW
    browser = webdriver.Chrome(service=chrome_service, options=chrome_options)
    wait = WebDriverWait(browser, 30)

    try:
        # Login
        browser.get('https://harbert31363.yardione.com/Account/Login?ReturnUrl=%2F')

        browser.find_element(By.CSS_SELECTOR, '#Username').send_keys(username)
        browser.find_element(By.ID, 'txtPassword').send_keys(password)
        wait.until(EC.element_to_be_clickable((By.NAME, 'button'))).click()

        # MFA
        time.sleep(120)
        mfa_code = get_mfa_code()
        print(f"MFA code: {mfa_code}")

        wait.until(EC.element_to_be_clickable((By.ID, 'verificationCode'))).send_keys(mfa_code)
        time.sleep(2)
        wait.until(EC.element_to_be_clickable((By.ID, 'verificationSubmit'))).click()
        time.sleep(5)

        # Navigate to Voyager
        element = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, 'Voyager 7s')))
        ActionChains(browser).move_to_element(element).click().perform()
        new_tab = browser.window_handles[-1]
        browser.switch_to.window(new_tab)
        time.sleep(3)

        browser.find_element(By.CSS_SELECTOR, '#cmdLogin').click()
        time.sleep(10)

        browser.find_element(By.LINK_TEXT, 'Roles').click()
        browser.find_element(By.LINK_TEXT, 'Residential Manager').click()
        time.sleep(2)

        # Work Orders report
        browser.get('https://www.yardiasptx11.com/33613harbert/pages/menu.aspx?sMenuSet=res')

        time.sleep(30)
        browser.find_element(By.ID, 'miFavorite').click()
        browser.find_element(By.LINK_TEXT, 'Work Orders').click()
        time.sleep(5)

        browser.switch_to.frame('filter')
        browser.find_element(By.ID, 'Clear_Button').click()

        browser.find_element(By.ID, 'PropertyLookUp_LookupCode').send_keys('res_main')

        # Open WOs (report type 1)
        Select(browser.find_element(By.ID, 'ReportType_DropDownList')).select_by_value('1')

        browser.find_element(By.ID, 'chkProblemDescription_CheckBox').click()
        browser.find_element(By.ID, 'chkTechnicianNotes_CheckBox').click()
        browser.find_element(By.ID, 'chkAccessNotes_CheckBox').click()
        browser.find_element(By.ID, 'chkFullDescription_CheckBox').click()
        browser.find_element(By.ID, 'chkCallerName_CheckBox').click()
        browser.find_element(By.ID, 'chkCallerPhone_CheckBox').click()
        browser.find_element(By.ID, 'chkCallerEmail_CheckBox').click()
        browser.find_element(By.ID, 'chkActualHours_CheckBox').click()
        time.sleep(5)

        browser.find_element(By.ID, 'Display_Button').click()
        time.sleep(10)

        # Click grand total drill-down (all properties)
        elem = browser.find_element(
            By.XPATH,
            "//a[contains(@href,'GrandTotal=Yes') and contains(@href,'ColName=Total')]"
        )
        elem.click()
        browser.switch_to.window(browser.window_handles[-1])

        btn = wait.until(EC.element_to_be_clickable((By.ID, 'Excel_Button')))
        ActionChains(browser).move_to_element(btn).click().perform()
        time.sleep(30)

        # Rename downloaded file
        files = glob.glob(os.path.join(DOWNLOAD_DIR, "*.xls"))
        if not files:
            raise FileNotFoundError("No .xls file found in download directory after Yardi export.")
        latest = max(files, key=os.path.getctime)
        dest = os.path.join(DOWNLOAD_DIR, "Open WO.xls")
        os.rename(latest, dest)
        print(f"Downloaded: {dest}")
        return dest

    finally:
        browser.quit()


# ---------- STEP 2: LOAD & CLEAN DATA ----------

def load_open_work_orders(xls_path: str) -> pd.DataFrame:
    """
    Load open WOs from the .xls Yardi export.
    Yardi exports have junk in rows 1-2; row 3 is the header.
    """
    df = pd.read_excel(xls_path, header=2)

    if "WO#" not in df.columns:
        df.rename(columns={df.columns[0]: "WO#"}, inplace=True)

    df = df[df["WO#"].notna()].copy()
    df = df[pd.to_numeric(df["WO#"], errors="coerce").notna()].copy()
    df.reset_index(drop=True, inplace=True)

    # Compute Days Open
    if "Call Date" in df.columns:
        df["Call Date Parsed"] = pd.to_datetime(df["Call Date"], errors="coerce")
        today = pd.Timestamp.today().normalize()
        df["Days Open"] = (today - df["Call Date Parsed"]).dt.days
    else:
        df["Days Open"] = None

    return df


def enrich_with_property_data(df: pd.DataFrame) -> pd.DataFrame:
    """Join property_name and tech_name onto df via the Property code, using PROPERTIES_MAP."""
    def get_prop(code, field):
        key = normalize_code(code)
        return PROPERTIES_MAP.get(key, {}).get(field, "")
    df["property_name"] = df["Property"].apply(lambda x: get_prop(x, "property_name"))
    df["tech_name"]     = df["Property"].apply(lambda x: get_prop(x, "tech_name"))
    return df


# ---------- STEP 3: AI TRIAGE ----------

def get_text_blob(row) -> str:
    parts = [
        str(row.get("Brief Desc.", "") or ""),
        str(row.get("Problem Description", "") or ""),
        str(row.get("Technician Notes", "") or ""),
    ]
    return " ".join(parts).lower()


def rule_based_priority(row):
    text = get_text_blob(row)
    for kw in EMERGENCY_KEYWORDS:
        if kw in text:
            return 1, "Hard emergency keyword detected (failsafe rule)."
    for kw in RISKY_WORDS:
        if kw in text:
            return 2, "Risk keyword detected (failsafe rule)."
    return None, None


def build_ai_prompt(row) -> str:
    days_open = row.get("Days Open", None)
    return f"""
You are helping a multifamily property manager triage maintenance work orders.

PRIORITY DEFINITIONS:
1 = Emergency (life/safety, habitability failure, active water intrusion, urgent electrical hazard)
2 = High (needs same-day service if at all possible)
3 = Medium (24-48 hours)
4 = Routine (standard work order, no serious risk if delayed briefly)

Consider days open: {days_open}

Work order details:
- WO#: {row.get("WO#", "")}
- Property: {row.get("Property", "")}
- Unit: {row.get("Unit", "")}
- Current Yardi Priority: {row.get("Priority", "")}
- Category: {row.get("Category", "")}
- Brief Desc: {row.get("Brief Desc.", "")}
- Problem Description: {row.get("Problem Description", "")}
- Technician Notes: {row.get("Technician Notes", "")}

Respond with ONLY valid JSON, no explanation text:
{{"priority": 2, "reason": "Short one-sentence explanation."}}
"""


def ai_priority_for_row(client, row) -> tuple:
    prompt = build_ai_prompt(row)
    try:
        response = client.chat.completions.create(
            model=OPENAI_MODEL,
            messages=[
                {"role": "system", "content": "You classify maintenance work orders."},
                {"role": "user", "content": prompt},
            ],
            max_tokens=150,
            timeout=20,
        )
        content = response.choices[0].message.content
        if not isinstance(content, str):
            content = "".join(
                p["text"] if isinstance(p, dict) and "text" in p else str(p)
                for p in content
            )
    except Exception as e:
        return 4, f"Error calling AI: {e!r}"

    try:
        data = json.loads(content)
        priority = max(1, min(4, int(data.get("priority", 4))))
        reason = str(data.get("reason", "") or "").strip() or "No explanation provided."
    except Exception:
        priority = 4
        reason = f"Failed to parse AI response: {content[:200]!r}"

    return priority, reason


def generate_ai_summary(client, row) -> str:
    brief    = str(row.get("Brief Desc.", "") or "")
    problem  = str(row.get("Problem Description", "") or "")
    full     = str(row.get("Full Description", "") or "")
    notes    = str(row.get("Technician Notes", "") or "")
    category = str(row.get("Category", "") or "")
    unit     = str(row.get("Unit", "") or "")

    prompt = f"""
Summarize this maintenance work order in ONE sentence, max 20 words.
Be specific and factual. No filler words. Start with the issue, not "resident reports".

Category: {category}
Unit: {unit}
Brief: {brief}
Problem: {problem[:300]}
Full description: {full[:300]}
Tech notes: {notes[:200]}

Respond with ONLY the summary sentence, nothing else.
"""
    try:
        response = client.chat.completions.create(
            model=OPENAI_MODEL,
            messages=[
                {"role": "system", "content": "You write concise maintenance work order summaries."},
                {"role": "user",   "content": prompt},
            ],
            max_tokens=60,
            timeout=15,
        )
        return response.choices[0].message.content.strip()
    except Exception:
        return brief or problem[:80] or "No description available."


def triage_work_orders(df: pd.DataFrame) -> pd.DataFrame:
    if not OPENAI_API_KEY:
        print("WARNING: OPENAI_API_KEY not set. Skipping AI triage, all set to priority 4.")
        df["AI Priority"] = 4
        df["AI Reason"] = "AI triage skipped — no API key."
        df["AI Summary"] = ""
        return df

    client = OpenAI(api_key=OPENAI_API_KEY)
    df = df.copy()
    ai_priorities, ai_reasons, ai_summaries = [], [], []
    limit = min(len(df), MAX_ROWS_FOR_AI)
    print(f"Running AI triage on {limit} work orders...")

    for i in range(limit):
        print(f"  Triaging {i+1}/{limit}...", flush=True)
        row = df.loc[i]
        p, reason = rule_based_priority(row)
        if p is not None:
            ai_priorities.append(p)
            ai_reasons.append(reason)
        else:
            p, reason = ai_priority_for_row(client, row)
            ai_priorities.append(p)
            ai_reasons.append(reason)
        ai_summaries.append(generate_ai_summary(client, row))

    # Any rows beyond limit: routine
    extra = len(df) - limit
    if extra > 0:
        ai_priorities.extend([4] * extra)
        ai_reasons.extend(["Not triaged by AI (row limit)."] * extra)
        ai_summaries.extend([""] * extra)

    df["AI Priority"] = ai_priorities
    df["AI Reason"] = ai_reasons
    df["AI Summary"] = ai_summaries
    return df


# ---------- STEP 4: SAVE TO XLSX ----------

def save_to_xlsx(df: pd.DataFrame, output_path: str, sheet_name: str):
    """Save the enriched dataframe to WorkOrders.xlsx."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Sort: AI Priority ASC, Days Open DESC
    sort_cols = [c for c in ["AI Priority", "Days Open"] if c in df.columns]
    if sort_cols:
        df = df.sort_values(
            by=sort_cols,
            ascending=[True, False],
            na_position="last"
        )

    with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Saved {len(df)} work orders to: {output_path} [{sheet_name}]")


# ---------- MAIN ----------

def main():
    global PROPERTIES_MAP

    print("=" * 50)
    print(f"harvest_sync.py — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 50)

    try:
        PROPERTIES_MAP = load_properties_map(PROPERTIES_XLSX_PATH)
        print(f"Loaded {len(PROPERTIES_MAP)} properties from {PROPERTIES_XLSX_PATH}.")
    except Exception as e:
        print(f"WARNING: Could not load properties file ({e}). Continuing without property enrichment.")
        PROPERTIES_MAP = {}

    # Step 1: Download from Yardi
    print("\n[1/3] Downloading open work orders from Yardi...")
    xls_path = download_open_wo()

    # Step 2: Load and clean
    print("\n[2/3] Loading and cleaning data...")
    df = load_open_work_orders(xls_path)
    df = enrich_with_property_data(df)
    print(f"  Loaded {len(df)} work orders.")

    # Step 3: AI triage
    print("\n[3/3] Running AI triage...")
    df = triage_work_orders(df)

    # Save enriched XLSX
    print("\nSaving enriched file...")
    save_to_xlsx(df, OUTPUT_PATH, OUTPUT_SHEET)

    print("\nDone. Run sync_workorders.py to sync this file to Railway.")


if __name__ == "__main__":
    main()
