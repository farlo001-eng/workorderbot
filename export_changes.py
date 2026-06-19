"""
export_changes.py
=================
Fetches pending work order changes from Railway and writes them to
WO_Changes.xlsx for Yardi writeback.

Run with: py export_changes.py
"""

import os
import json
import requests
import openpyxl
from datetime import datetime
from urllib.parse import quote
from dotenv import load_dotenv

load_dotenv()

RAILWAY_URL  = os.getenv("RAILWAY_URL", "").rstrip("/")
SYNC_API_KEY = os.getenv("SYNC_API_KEY", "harvest-workorder-2026-secure-key")
OUTPUT_PATH  = r"I:\PycharmProjects\WorkOrderBot\WO_Changes.xlsx"

LOCAL_PHOTOS_DIR = r"G:\Harvest Apartment Management\Work Orders\Photos"

PENDING_SHEET   = "Pending"
COMPLETED_SHEET = "Completed"

HEADERS = [
    "Change ID", "WO#", "Property", "Unit", "Brief Desc",
    "Field Changed", "Old Value", "New Value",
    "Changed By", "Changed At", "Assigned Tech", "Local Photo Paths"
]


def fetch_pending_changes():
    url = f"{RAILWAY_URL}/api/changes/export"
    headers = {"X-API-Key": SYNC_API_KEY}
    resp = requests.get(url, headers=headers, timeout=30)
    resp.raise_for_status()
    return resp.json()


def mark_completed(ids: list):
    if not ids:
        return
    url = f"{RAILWAY_URL}/api/changes/mark_complete"
    headers = {
        "Content-Type": "application/json",
        "X-API-Key": SYNC_API_KEY,
    }
    resp = requests.post(url, headers=headers, json={"ids": ids}, timeout=30)
    resp.raise_for_status()
    return resp.json()


def load_or_create_workbook(path: str) -> openpyxl.Workbook:
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    # Create Pending sheet
    ws_pending = wb.active
    ws_pending.title = PENDING_SHEET
    ws_pending.append(HEADERS)
    # Create Completed sheet
    ws_completed = wb.create_sheet(COMPLETED_SHEET)
    ws_completed.append(HEADERS)
    return wb


def get_or_create_sheet(wb: openpyxl.Workbook, name: str) -> openpyxl.worksheet.worksheet.Worksheet:
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.append(HEADERS)
    return ws


def download_photos(changes: list) -> dict:
    """
    For any pending change rows where field == 'photos',
    download the photo files from Railway to LOCAL_PHOTOS_DIR.
    Skips files that already exist locally.
    Returns a dict mapping original Railway path -> local file path.
    """
    os.makedirs(LOCAL_PHOTOS_DIR, exist_ok=True)

    downloaded = {}
    headers = {"X-API-Key": SYNC_API_KEY}

    for change in changes:
        if change.get("field") != "photos":
            continue

        # new_value is a JSON array of photo paths like ["photos/WO#123_....jpg", ...]
        try:
            photo_paths = json.loads(change.get("new_value", "[]"))
            if not isinstance(photo_paths, list):
                continue
        except Exception:
            continue

        for path in photo_paths:
            if not path or path in downloaded:
                continue

            # Extract filename from path (e.g. "photos/WO#30583_20260619_143022123.jpg")
            filename = os.path.basename(path)
            local_path = os.path.join(LOCAL_PHOTOS_DIR, filename)

            # Skip if already downloaded
            if os.path.exists(local_path):
                downloaded[path] = local_path
                print(f"  Photo already exists locally: {filename}")
                continue

            # Download from Railway — URL-encode the filename to handle the # character
            encoded_filename = quote(filename, safe='')
            url = f"{RAILWAY_URL}/photos/{encoded_filename}"

            try:
                resp = requests.get(url, headers=headers, timeout=30, stream=True)
                resp.raise_for_status()
                with open(local_path, "wb") as f:
                    for chunk in resp.iter_content(chunk_size=8192):
                        f.write(chunk)
                downloaded[path] = local_path
                print(f"  Downloaded photo: {filename} → {local_path}")
            except requests.exceptions.RequestException as e:
                print(f"  WARNING: Could not download photo {filename}: {e}")

    return downloaded


def change_to_row(change: dict, downloaded: dict = None) -> list:
    local_paths = ""
    if change.get("field") == "photos" and downloaded:
        try:
            photo_paths = json.loads(change.get("new_value", "[]"))
            if isinstance(photo_paths, list):
                local_paths = "; ".join(
                    downloaded.get(p, "") for p in photo_paths if downloaded.get(p)
                )
        except Exception:
            pass

    return [
        change.get("id"),
        change.get("wo_id"),
        change.get("property_name", ""),
        change.get("unit_number", ""),
        change.get("brief_desc", ""),
        change.get("field"),
        change.get("old_value", ""),
        change.get("new_value", ""),
        change.get("changed_by", ""),
        change.get("changed_at", ""),
        change.get("tech_name", ""),
        local_paths,
    ]


def main():
    if not RAILWAY_URL:
        print("ERROR: RAILWAY_URL not set in .env")
        return

    print(f"Fetching pending changes from {RAILWAY_URL}...")
    try:
        changes = fetch_pending_changes()
    except requests.exceptions.RequestException as e:
        print(f"ERROR: Could not fetch changes — {e}")
        return

    if not changes:
        print("No pending changes found.")
        return

    print(f"Found {len(changes)} pending change(s).")

    # Download any photos referenced in pending changes
    downloaded = {}
    if any(c.get("field") == "photos" for c in changes):
        print(f"\nDownloading photos to {LOCAL_PHOTOS_DIR}...")
        downloaded = download_photos(changes)
        print(f"Downloaded {len(downloaded)} photo(s).")
    else:
        print("No photo changes to download.")

    wb = load_or_create_workbook(OUTPUT_PATH)
    ws_pending   = get_or_create_sheet(wb, PENDING_SHEET)
    ws_completed = get_or_create_sheet(wb, COMPLETED_SHEET)

    # Move existing Pending rows to Completed sheet first
    pending_rows = list(ws_pending.iter_rows(min_row=2, values_only=True))
    if pending_rows:
        print(f"Moving {len(pending_rows)} existing pending row(s) to Completed...")
        for row in pending_rows:
            ws_completed.append(list(row))
        # Clear pending sheet (keep header)
        for row in ws_pending.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        # Delete the blank rows
        ws_pending.delete_rows(2, ws_pending.max_row)

    # Write new pending changes to Pending sheet
    ids_to_mark = []
    for change in changes:
        ws_pending.append(change_to_row(change, downloaded))
        ids_to_mark.append(change["id"])

    wb.save(OUTPUT_PATH)
    print(f"Saved {len(changes)} change(s) to: {OUTPUT_PATH} [{PENDING_SHEET}]")

    # Mark as completed in Railway
    print("Marking changes as completed in Railway...")
    try:
        result = mark_completed(ids_to_mark)
        print(f"Marked {result.get('marked', '?')} change(s) as completed.")
    except requests.exceptions.RequestException as e:
        print(f"WARNING: Could not mark changes as completed — {e}")
        print("Changes were written to Excel but remain 'pending' in Railway.")

    print("Done.")


if __name__ == "__main__":
    main()
